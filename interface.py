'''
Script to request user input on end date, time horizon,
and filter; returns package and benchmark performances
'''

import os
import time
import datetime
import numpy as np
import pandas as pd
from plotter import plot_chart
from calendar import monthrange
from create_spreadsheet import make_workbook
from overlayPNGs import overlay,overlay_heatmap
from get_index_name import get_index_name
from get_index import get_index
import openpyxl
import shutil
import glob
import xlrd
import csv
from tqdm import tqdm
import pdb


def fetch_packages(start_date,end_date):
    '''
    Get packages from database that correspond to end date
    and forecast length
    '''

    short_month_names = {'01': 'Jan','02': 'Feb','03': 'Mar','04': 'Apr','05': 'May','06': 'Jun',
                        '07': 'Jul','08': 'Aug','09': 'Sep','10': 'Oct','11': 'Nov','12': 'Dec'}
    files = []
    search_date_info = start_date.split('/')
    if len(search_date_info[0]) == 1:
        search_date_info[0] = '0' + search_date_info[0]
    if len(search_date_info[1]) == 1:
        search_date_info[1] = '0' + search_date_info[1]
    search_date_info[2] = '20' + search_date_info[2]
    search_date = search_date_info[1] + '_' + short_month_names[search_date_info[0]] + '_' + search_date_info[2]
    #Imput data folder should include date_to_get
    path = 'Input/'+short_month_names[search_date_info[0]]+'_'+search_date_info[2]+'/'+search_date+'/*.xls'
    files = glob.glob(path)

    return files

def load_heatmap(data_file,forecast_length):
    '''
    Function to return dictionary with heat map information

    Arguments:
    ----------
        data_file -- data forecast file (must be in .xlsx format)
        forecast_length -- length of forecast to make chart for (ex. '3', '7', '14'...)
    Returns:
    -------
        package_dict -- dictionary with information on heatmap boxes
    '''

    sp_keys = ['_Stocks_SP500_',
                '_top_5_SP500_',
                '_top_10_SP500_',
                '_top_15_SP500_',
                '_top_20_SP500_',
                '_top_40_SP500_']

    sp_chart = False
    for key in sp_keys:
        if key in data_file:
            sp_chart = True
            break

    #Set column limits based on which forecast heatmap to make
    if forecast_length == '3':
        start_sheet = 0
        start_col = 1
        end_col = 5
    elif forecast_length == '7':
        start_sheet = 0
        start_col = 7
        end_col = 11
    elif forecast_length == '14':
        start_sheet = 0
        start_col = 13
        end_col = 17
    elif forecast_length == '30':
        start_sheet = 1
        start_col = 1
        end_col = 5
    elif forecast_length == '90':
        start_sheet = 1
        start_col = 7
        end_col = 11
    elif forecast_length == '365':
        start_sheet = 1
        start_col = 13
        end_col = 17


    #S&P Info for additional spreadsheet column
    sp_info = None

    #Initialize return dictionary
    package_dict = {}

    #Open Workbook
    book = xlrd.open_workbook(data_file,formatting_info=True)
    sheets = book.sheet_names()
    sheet = book.sheet_by_index(start_sheet)
    rows, cols = sheet.nrows, sheet.ncols
    if cols < 6:
        return package_dict, sp_return
    #Loop through rows in sheet
    try:
    # if True:
        for row in range(5,rows-2,3):
            #Loop through correct columns
            for col in range(start_col,end_col+1):
                thecell = sheet.cell(row,col)
                xfx = sheet.cell_xf_index(row,col)
                xf = book.xf_list[xfx]
                bgx = xf.background.pattern_colour_index
                if thecell.value == '':
                    pass
                else:
                    package_dict[thecell.value] = [sheet.cell(row+1,col).value,sheet.cell(row+2,col).value,bgx]
                if (bgx == 11 or bgx == 42) and sp_chart and thecell.value == '^S&P500':
                    sp_info = 'long'
                elif (bgx == 10 or bgx == 29) and sp_chart and thecell.value == '^S&P500':
                    sp_info = 'short'


    except Exception as e:
        # print(e)
        pass
    return package_dict, sp_info



def fast_get_ROR(horizon,asset_name,start_i,end_i):
    if asset_name == 'CME_IE1':
        asset_name = 'CME_CL1'
    try:
    # if True:
        # for array in all_frames:
        for key, value in all_frames.items():
            if end_i == key:
                end_price = value[asset_name]
                break

        for key, value in all_frames.items():
            if get_days_past(start_i,1) == key:
                start_price = value[asset_name]
                break

        return round(end_price / start_price - 1,5)

    except Exception as e:
        # print(e)
        return 0

def get_day_of_week(date):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int('20'+date_info[2])
    #Get current day of the week
    current_weekday = datetime.datetime(year,month,day).weekday()
    return current_weekday

def get_days_in_month(date):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int('20'+date_info[2])
    #Get number of days in month

    return monthrange(year,month)[1]

def get_days_past(date,num):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int('20'+date_info[2])
    previous_date = (datetime.date(year, month, day) - datetime.timedelta(num)).isoformat()
    previous_date_info = previous_date.split('-')
    new_day = previous_date_info[2]
    if new_day[0] == '0':
        new_day = new_day[1]
    new_month = previous_date_info[1]
    if new_month[0] == '0':
        new_month = new_month[1]
    new_year = previous_date_info[0][-2:]
    return str(new_month)+'/'+str(new_day)+'/'+str(new_year)

def get_1_month(date):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int(date_info[2])
    new_day = day
    new_year = year
    if month == 1:
        new_month = 12
        new_year = year - 1
    else:
        new_month = month - 1

    max_days = get_days_in_month(str(new_month)+'/'+str(new_day)+'/'+str(new_year))
    if new_day > max_days:
        new_day = max_days

    if get_day_of_week(str(new_month)+'/'+str(new_day)+'/'+str(new_year)) == 5:
        return get_days_past(str(new_month)+'/'+str(new_day)+'/'+str(new_year),1)
    return str(new_month)+'/'+str(new_day)+'/'+str(new_year)

def get_3_months(date):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int(date_info[2])
    new_day = day
    new_year = year
    if month == 1:
        new_month = 10
        new_year = year - 1
    elif month == 2:
        new_month = 11
        new_year = year - 1
    elif month == 3:
        new_month = 12
        new_year = year - 1
    else:
        new_month = month - 3

    max_days = get_days_in_month(str(new_month)+'/'+str(new_day)+'/'+str(new_year))
    if new_day > max_days:
        new_day = max_days

    if get_day_of_week(str(new_month)+'/'+str(new_day)+'/'+str(new_year)) == 5:
        return get_days_past(str(new_month)+'/'+str(new_day)+'/'+str(new_year),1)
    return str(new_month)+'/'+str(new_day)+'/'+str(new_year)

def get_1_year(date):
    date_info = date.split('/')
    day = int(date_info[1])
    month = int(date_info[0])
    year = int(date_info[2])
    new_day = day
    new_month = month
    new_year = year - 1

    max_days = get_days_in_month(str(new_month)+'/'+str(new_day)+'/'+str(new_year))
    if new_day > max_days:
        new_day = max_days

    if get_day_of_week(str(new_month)+'/'+str(new_day)+'/'+str(new_year)) == 5:
        return get_days_past(str(new_month)+'/'+str(new_day)+'/'+str(new_year),1)
    return str(new_month)+'/'+str(new_day)+'/'+str(new_year)


#Load data into data frame
end_date = input("\nEnter the date you would like forecast information for (format: 3/1/19):\n\n")
df = pd.read_csv('datcommodcurr.csv',header=1,index_col=0)


end_date_info = end_date.split('/')
day = end_date_info[1]
month = end_date_info[0]
year = end_date_info[2]
day_is_sunday = False
day_is_monday = False
day_is_tuesday = False
day_is_wednesday = False
day_is_thursday = False
day_is_friday = False
day_is_saturday = False

all_frames = {}

#Get current day of the week
current_weekday = get_day_of_week(end_date)
if current_weekday == 6:
    day_is_sunday = True
    end_dates = [end_date,
                get_days_past(end_date,1),
                get_days_past(end_date,2),
                get_days_past(end_date,3)]

    end_indexes = [df.index.get_loc(end_date),
                    df.index.get_loc(get_days_past(end_date,1)),
                    df.index.get_loc(get_days_past(end_date,2)),
                    df.index.get_loc(get_days_past(end_date,3))]

    start_date0 = [df.index[end_indexes[0]-3],
                    df.index[end_indexes[1]-3],
                    df.index[end_indexes[2]-3],
                    df.index[end_indexes[3]-3]]

    start_date1 = [df.index[end_indexes[0]-7],
                    df.index[end_indexes[2]-7],
                    df.index[end_indexes[3]-7]]

    start_date2 = [df.index[end_indexes[0]-14],
                    df.index[end_indexes[2]-14],
                    df.index[end_indexes[3]-14]]

    start_date3 = [get_1_month(end_dates[0]),
                    get_1_month(end_dates[1]),
                    get_1_month(end_dates[2]),
                    get_1_month(end_dates[3])]

    start_date4 = [get_3_months(end_dates[0]),
                    get_3_months(end_dates[1]),
                    get_3_months(end_dates[2]),
                    get_3_months(end_dates[3])]

    start_date5 = [get_1_year(end_dates[0]),
                    get_1_year(end_dates[1]),
                    get_1_year(end_dates[2]),
                    get_1_year(end_dates[3])]

elif current_weekday == 0:
    day_is_monday = True
    end_dates = [end_date]
    end_indexes = [df.index.get_loc(end_date)]

    start_date0 = [df.index[end_indexes[0]-3]]
    start_date1 = [df.index[end_indexes[0]-7]]
    start_date2 = [df.index[end_indexes[0]-14]]
    start_date3 = [get_1_month(end_dates[0])]
    start_date4 = [get_3_months(end_dates[0])]
    start_date5 = [get_1_year(end_dates[0])]

elif current_weekday == 1:
    day_is_tuesday = True
    end_dates = [end_date]
    end_indexes = [df.index.get_loc(end_date)]

    start_date0 = [df.index[end_indexes[0]-4]]
    start_date1 = [df.index[end_indexes[0]-7]]
    start_date2 = [df.index[end_indexes[0]-14]]
    start_date3 = [get_1_month(end_dates[0])]
    start_date4 = [get_3_months(end_dates[0])]
    start_date5 = [get_1_year(end_dates[0])]

elif current_weekday == 2:
    day_is_wednesday = True
    end_dates = [end_date]
    end_indexes = [df.index.get_loc(end_date)]

    start_date0 = [df.index[end_indexes[0]-3]]
    start_date1 = [df.index[end_indexes[0]-7]]
    start_date2 = [df.index[end_indexes[0]-14]]
    start_date3 = [get_1_month(end_dates[0])]
    start_date4 = [get_3_months(end_dates[0])]
    start_date5 = [get_1_year(end_dates[0])]

elif current_weekday == 3:
    day_is_thursday = True
    end_dates = [end_date]
    end_indexes = [df.index.get_loc(end_date)]

    start_date0 = [df.index[end_indexes[0]-3]]
    start_date1 = [df.index[end_indexes[0]-7]]
    start_date2 = [df.index[end_indexes[0]-14]]
    start_date3 = [get_1_month(end_dates[0])]
    start_date4 = [get_3_months(end_dates[0])]
    start_date5 = [get_1_year(end_dates[0])]

elif current_weekday == 4:
    day_is_friday = True
    end_dates = [end_date]
    end_indexes = [df.index.get_loc(end_date)]

    start_date0 = [df.index[end_indexes[0]-3]]
    start_date1 = [df.index[end_indexes[0]-7]]
    start_date2 = [df.index[end_indexes[0]-14]]
    start_date3 = [get_1_month(end_dates[0])]
    start_date4 = [get_3_months(end_dates[0])]
    start_date5 = [get_1_year(end_dates[0])]


print(start_date0)
print(start_date1)
print(start_date2)
print(start_date3)
print(start_date4)
print(start_date5)

start_date0_index = []
start_date1_index = []
start_date2_index = []
start_date3_index = []
start_date4_index = []
start_date5_index = []

for x in range(len(start_date0)):
    start_date0_index.append(df.index.get_loc(start_date0[x]))
for x in range(len(start_date1)):
    start_date1_index.append(df.index.get_loc(start_date1[x]))
for x in range(len(start_date2)):
    start_date2_index.append(df.index.get_loc(start_date2[x]))
for x in range(len(start_date3)):
    start_date3_index.append(df.index.get_loc(start_date3[x]))
for x in range(len(start_date4)):
    start_date4_index.append(df.index.get_loc(start_date4[x]))
for x in range(len(start_date5)):
    start_date5_index.append(df.index.get_loc(start_date5[x]))

start_date0_ROR = []
start_date1_ROR = []
start_date2_ROR = []
start_date3_ROR = []
start_date4_ROR = []
start_date5_ROR = []

for x in range(len(start_date0)):
    start_date0_ROR.append(df.index[start_date0_index[x] - 1])
for x in range(len(start_date1)):
    start_date1_ROR.append(df.index[start_date1_index[x] - 1])
for x in range(len(start_date2)):
    start_date2_ROR.append(df.index[start_date2_index[x] - 1])
for x in range(len(start_date3)):
    start_date3_ROR.append(df.index[start_date3_index[x] - 1])
for x in range(len(start_date4)):
    start_date4_ROR.append(df.index[start_date4_index[x] - 1])
for x in range(len(start_date5)):
    start_date5_ROR.append(df.index[start_date5_index[x] - 1])

array_0 = []
array_1 = []
array_2 = []
array_3 = []
array_4 = []
array_5 = []

for x in range(len(start_date0_ROR)):
    array_0.append(df.loc[start_date0_ROR[x]])
    all_frames[start_date0_ROR[x]] = (array_0[-1])
for x in range(len(start_date1_ROR)):
    array_1.append(df.loc[start_date1_ROR[x]])
    all_frames[start_date1_ROR[x]] = (array_1[-1])
for x in range(len(start_date2_ROR)):
    array_2.append(df.loc[start_date2_ROR[x]])
    all_frames[start_date2_ROR[x]] = (array_2[-1])
for x in range(len(start_date3_ROR)):
    array_3.append(df.loc[start_date3_ROR[x]])
    all_frames[start_date3_ROR[x]] = (array_3[-1])
for x in range(len(start_date4_ROR)):
    array_4.append(df.loc[start_date4_ROR[x]])
    all_frames[start_date4_ROR[x]] = (array_4[-1])
for x in range(len(start_date5_ROR)):
    array_5.append(df.loc[start_date5_ROR[x]])
    all_frames[start_date5_ROR[x]] = (array_5[-1])

end_array = []
for end_date in end_dates:
    end_array.append(df.loc[end_date])
    all_frames[end_date] = end_array[-1]



#Get user input
# startTime = time.time()

folder_date_info = end_dates[0].split('/')
if len(folder_date_info[0]) == 1:
    folder_date_info[0] = '0' + folder_date_info[0]
if len(folder_date_info[1]) == 1:
    folder_date_info[1] = '0' + folder_date_info[1]
folder_date_info[2] = '20' + folder_date_info[2]
output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]
print(output_date)
# try:
#     os.mkdir('Output-'+output_date+'/')
#     output_folder = 'Output-'+output_date+'/'
# except:
#     output_folder = 'Output-'+output_date+'/'


short_month_names = {'1': 'Jan','2': 'Feb','3': 'Mar','4': 'Apr','5': 'May','6': 'Jun',
                '7': 'Jul','8': 'Aug','9': 'Sep','10': 'Oct','11': 'Nov','12': 'Dec'}

month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

horizon_to_date = {'365':start_date5,'90':start_date4,'30':start_date3,'14':start_date2,'7':start_date1,'3':start_date0}
horizons = [
            '3',
            '7',
            '14',
            '30',
            '90',
            '365'
            ]
horizon_Mat = []

performance_dict = {'3':{},'7':{},'14':{},'30':{},'90':{},'365':{}}
best_currency = 0
for horizon in horizons:
    try:
    # if True:
        print(horizon)
        current_horizon_Mat = []
        start_date_list = horizon_to_date[horizon]

        count_i = 0
        for z, start_date in enumerate(start_date_list):
            end_date = end_dates[z]
            if horizon == '7' and count_i == 2 or \
                horizon == '14' and count_i == 2:
                end_date = end_dates[z+1]


            files2 = []
            get_more_files = False

            files = fetch_packages(start_date,end_date)

            all_files2 = []
            exempt_pack = ['XxXxXx']
            if len(files) < 400:
                for item in files:
                    for month_name in month_names:
                        if month_name in item:
                            exempt_pack.append(item.split(month_name)[0][:-4])
                            break

                    if not exempt_pack[-1] in item:
                        exempt_pack.append(item)

                print("Need to get more files")
                get_more_files = True
                look_back = 1
                while get_more_files:
                    all_files2 = fetch_packages(get_days_past(start_date,look_back),end_date)
                    look_back += 1
                    if len(all_files2) > 200:
                        get_more_files = False

                for g in range(len(all_files2)):
                    for h in range(len(exempt_pack)):
                        if not exempt_pack[h] in all_files2[g]:
                            files2.append(all_files2[g])
                            break


            all_files = [files,files2]
            k_index = 0
            for files in all_files:

                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                k_index += 1
                for file in tqdm(files,desc="Processing Files for "+str(horizon)+" day forecasts v"+str(k_index)):
                    # print(file)
                    if 'CUSTOM' in file or \
                        'wk1' in file or \
                        'best' in file or \
                        'macro' in file:
                        continue

                    long_short_top_list = {'Long':[],'Short':[]}
                    # print(file)

                    sp_return = None
                    try:
                    # if True:
                        long_top5_asset_list_ROR = []
                        short_top5_asset_list_ROR = []
                        long_top10_asset_list_ROR = []
                        short_top10_asset_list_ROR = []
                        long_top20_asset_list_ROR = []
                        short_top20_asset_list_ROR = []
                        long_hit_ratio_tracker = []
                        short_hit_ratio_tracker = []
                        benchmark_name = get_index(file)
                        # benchmark_ROR = fast_get_ROR(int(horizon),benchmark_name,count_i,count_j)
                        benchmark_ROR = fast_get_ROR(int(horizon),benchmark_name,start_date,end_date)
                        file_dict, sp_info = load_heatmap(file,horizon)
                        counter = 0
                        info_list = []
                        long_zero = 0
                        short_zero = 0

                        if sp_info is not None:
                            if sp_info == 'short':
                                sp_return = -fast_get_ROR(int(horizon),'^S&P500',start_date,end_date)
                            else:
                                sp_return = fast_get_ROR(int(horizon),'^S&P500',start_date,end_date)

                        for key, value in file_dict.items():

                            if 'currencies' in file or 'Currencies' in file:
                                # ror = fast_get_ROR(int(horizon),key,count_i,count_j)
                                ror = fast_get_ROR(int(horizon),key,start_date,end_date)
                                if value[-1] == 11 or value[-1] == 42:
                                    if ror == 0:
                                        long_zero += 1
                                        short_zero += 1
                                    if ror > 0:
                                        long_hit_ratio_tracker.append(1)
                                        short_hit_ratio_tracker.append(1)
                                    else:
                                        long_hit_ratio_tracker.append(0)
                                        short_hit_ratio_tracker.append(0)

                                elif value[-1] == 10 or value[-1] == 29:
                                    if ror == 0:
                                        long_zero += 1
                                        short_zero += 1
                                    if ror < 0:
                                        ror = -ror
                                        long_hit_ratio_tracker.append(1)
                                        short_hit_ratio_tracker.append(1)
                                    else:
                                        long_hit_ratio_tracker.append(0)
                                        short_hit_ratio_tracker.append(0)

                                long_top20_asset_list_ROR.append(ror)
                                long_top10_asset_list_ROR.append(ror)
                                long_top5_asset_list_ROR.append(ror)
                                short_top20_asset_list_ROR.append(ror)
                                short_top10_asset_list_ROR.append(ror)
                                short_top5_asset_list_ROR.append(ror)

                            else:
                                if value[-1] == 11 or value[-1] == 42:
                                    if counter < 20:
                                        # ror = fast_get_ROR(int(horizon),key,count_i,count_j)
                                        ror = fast_get_ROR(int(horizon),key,start_date,end_date)
                                        long_top20_asset_list_ROR.append(ror)
                                        if counter < 10:
                                            long_top10_asset_list_ROR.append(ror)
                                            long_short_top_list['Long'].append([key, ror])
                                            if ror == 0:
                                                long_zero += 1
                                            if ror > 0:
                                                long_hit_ratio_tracker.append(1)
                                            else:
                                                long_hit_ratio_tracker.append(0)
                                            if counter < 5:
                                                long_top5_asset_list_ROR.append(ror)

                                if value[-1] == 10 or value[-1] == 29:
                                    if counter >= (len(file_dict) - 20):
                                        # ror = fast_get_ROR(int(horizon),key,count_i,count_j)
                                        ror = fast_get_ROR(int(horizon),key,start_date,end_date)
                                        short_top20_asset_list_ROR.append(-ror)
                                        if counter >= (len(file_dict) - 10):
                                            short_top10_asset_list_ROR.append(-ror)
                                            long_short_top_list['Short'].append([key, -ror])
                                            if ror == 0:
                                                short_zero += 1
                                            if ror < 0:
                                                short_hit_ratio_tracker.append(1)
                                            else:
                                                short_hit_ratio_tracker.append(0)
                                            if counter >= (len(file_dict) - 5):
                                                short_top5_asset_list_ROR.append(-ror)

                            counter += 1

                        long_top20_total_ROR = round(np.mean(long_top20_asset_list_ROR),5)
                        if np.isnan(long_top20_total_ROR):
                            long_top20_total_ROR = 0

                        long_top10_total_ROR = round(np.mean(long_top10_asset_list_ROR),5)
                        if np.isnan(long_top10_total_ROR):
                            long_top10_total_ROR = 0

                        long_top5_total_ROR = round(np.mean(long_top5_asset_list_ROR),5)
                        if np.isnan(long_top5_total_ROR):
                            long_top5_total_ROR = 0

                        short_top20_total_ROR = round(np.mean(short_top20_asset_list_ROR),5)
                        if np.isnan(short_top20_total_ROR):
                            short_top20_total_ROR = 0

                        short_top10_total_ROR = round(np.mean(short_top10_asset_list_ROR),5)
                        if np.isnan(short_top10_total_ROR):
                            short_top10_total_ROR = 0

                        short_top5_total_ROR = round(np.mean(short_top5_asset_list_ROR),5)
                        if np.isnan(short_top5_total_ROR):
                            short_top5_total_ROR = 0


                        mark = '-'
                        long_hit_ratio = round(np.mean(long_hit_ratio_tracker),5)
                        short_hit_ratio = round(np.mean(short_hit_ratio_tracker),5)
                        if 'currencies' in file or 'Currencies' in file and len(long_top20_asset_list_ROR) > 10:
                            if long_hit_ratio > .5 and long_hit_ratio > best_currency:
                                best_currency = long_hit_ratio
                        if np.isnan(long_hit_ratio):
                            long_hit_ratio = 0
                        if np.isnan(short_hit_ratio):
                            short_hit_ratio = 0

                        if True:
                            if len(long_top20_asset_list_ROR) < 5:
                                long_top20_total_ROR = 0
                                long_top10_total_ROR = 0
                                long_top5_total_ROR = 0
                                long_hit_ratio = 0
                            elif len(long_top20_asset_list_ROR) < 10:
                                long_top20_total_ROR = 0
                                long_top10_total_ROR = 0
                                long_hit_ratio = 0
                            elif len(long_top20_asset_list_ROR) < 20:
                                long_top20_total_ROR = 0

                            if len(short_top20_asset_list_ROR) < 5:
                                short_top20_total_ROR = 0
                                short_top10_total_ROR = 0
                                short_top5_total_ROR = 0
                                short_hit_ratio = 0
                            elif len(short_top20_asset_list_ROR) < 10:
                                short_top20_total_ROR = 0
                                short_top10_total_ROR = 0
                                short_hit_ratio = 0
                            elif len(short_top20_asset_list_ROR) < 20:
                                short_top20_total_ROR = 0



                        # if 'commodities' in file or 'Commodities' in file:
                        if False:
                            benchmark = 'N/A'
                            benchmark_ROR = 0
                        else:
                            benchmark = get_index_name(benchmark_name)
                        info_list = ['Link',
                                    file,
                                    start_date,
                                    end_date,
                                    long_top5_total_ROR,
                                    long_top10_total_ROR,
                                    long_top20_total_ROR,
                                    long_hit_ratio,
                                    long_zero,
                                    short_top5_total_ROR,
                                    short_top10_total_ROR,
                                    short_top20_total_ROR,
                                    short_hit_ratio,
                                    short_zero,
                                    sp_return,
                                    benchmark,
                                    benchmark_ROR,
                                    mark,
                                    mark,
                                    mark]


                        # print(info_list)
                        # print('\n')

                        # index += 1
                    except Exception as e:
                        #print(e)
                        # index += 1
                        continue

                    long_short_top_list['Long'].append([get_index_name(benchmark_name),benchmark_ROR])
                    long_short_top_list['Short'].append([get_index_name(benchmark_name),benchmark_ROR])
                    current_horizon_Mat.append(info_list)

                    performance_dict[horizon][file] = long_short_top_list
            count_i += 1
        horizon_Mat.append(current_horizon_Mat)

    except Exception as e:
        # print(e)
        horizon_Mat.append([])

#pdb.set_trace()
make_workbook(output_date,horizon_Mat,performance_dict,best_currency)

# 'Input/Mar_2019/12_Mar_2019/IKForecast_top_10_big_Israel_12_Mar_2019.xls'



# print(time.time()-startTime)













#
