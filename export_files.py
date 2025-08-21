import os
import csv
import time
import shutil
import numpy as np
from plotter import plot_chart
from create_spreadsheet import make_workbook
from make_heatmap import start_process, print_heat_map, make_heatmap, get_tables
from overlayPNGs import overlay,overlay_heatmap
from get_index import get_index
from generate_xml import generate_xml
# from tabulate import tabulate
import pandas as pd
import openpyxl
import xlrd
import glob


def export_files():
    end_date = input("\nEnter the date you would like forecast information for (format: 3/1/20):\n\n")

    end_date_info = end_date.split('/')
    folder_date_info = end_date.split('/')
    if len(folder_date_info[0]) == 1:
        folder_date_info[0] = '0' + folder_date_info[0]
    if len(folder_date_info[1]) == 1:
        folder_date_info[1] = '0' + folder_date_info[1]
    folder_date_info[2] = '20' + folder_date_info[2]
    output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]
    if not os.path.isdir(output_date+'/'):
        os.makedirs(output_date+'/')
    else:
        shutil.rmtree(output_date+'/')
        os.makedirs(output_date+'/')


    short_month_names = {'1': 'Jan','2': 'Feb','3': 'Mar','4': 'Apr','5': 'May','6': 'Jun',
                    '7': 'Jul','8': 'Aug','9': 'Sep','10': 'Oct','11': 'Nov','12': 'Dec'}

    month_names = {'1': 'January','2': 'February','3': 'March','4': 'April','5': 'May','6': 'June',
                    '7': 'July','8': 'August','9': 'September','10': 'October','11': 'November','12': 'December'}

    horizons = ['3','7','14','30','90','365']

    reverse_month_names = {'Jan':'01','Feb': '02','Mar': '03','Apr': '04','May': '05','Jun': '06',
                    'Jul': '07','Aug': '08','Sep': '09','Oct': '10','Nov': '11','Dec': '12'}


    all_output_folders = []
    true_dates = {}
    the_sheet_names = {}

    book = xlrd.open_workbook(output_date+'.xlsx',on_demand=True)
    abs_date = output_date
    sheets = book.sheet_names()
    index = 0

    # Loop through forecast lengths/sheets
    for sheet in sheets:
        if index == 0:
            index += 1
            continue
        sheet = book.sheet_by_index(index)
        rows, cols = sheet.nrows, sheet.ncols

        reached_end = False

        long_files = []
        short_files = []
        long_short_files = []
        for row in range(3,rows):
            if index == 1:
                horizon_info = '3 days'
            elif index == 2:
                horizon_info = '7 days'
            elif index == 3:
                horizon_info = '14 days'
            elif index == 4:
                horizon_info = '30 days'
            elif index == 5:
                horizon_info = '90 days'
            elif index == 6:
                horizon_info = '1 year'

            # Get Long positions
            if sheet.cell(row,18).value == 'x' or \
                sheet.cell(row,18).value == 'X' or \
                sheet.cell(row,18).value == 'xs' or \
                sheet.cell(row,18).value == 'XS':

                long_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = long_files[-1].replace('\\','/').split('/')[-1].split('.')[0]


                file_info = long_files[-1].replace('\\','/').split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date

                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,18).value == 'xs' or \
                    sheet.cell(row,18).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' long' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' long' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.replace('\\','/').split('/')[-1]] = (long_files[-1].split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

            # Get Short positions
            if sheet.cell(row,19).value == 'x' or \
                sheet.cell(row,19).value == 'X' or \
                sheet.cell(row,19).value == 'xs' or \
                sheet.cell(row,19).value == 'XS':

                short_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = short_files[-1].split('/')[-1].split('.')[0]
                file_info = short_files[-1].split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date
                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,19).value == 'xs' or \
                    sheet.cell(row,19).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.replace('\\','/').split('/')[-1]] = (short_files[-1].replace('\\','/').split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

            # Get Long and Short positions
            if sheet.cell(row,20).value == 'x' or \
                sheet.cell(row,20).value == 'X' or \
                sheet.cell(row,20).value == 'xs' or \
                sheet.cell(row,20).value == 'XS':

                long_short_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = long_short_files[-1].split('/')[-1].split('.')[0]
                file_info = long_short_files[-1].split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date
                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,20).value == 'xs' or \
                    sheet.cell(row,20).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' long+short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + " " + horizon_info + ' long+short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.split('/')[-1]] = (long_short_files[-1].split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

            # Get Long 20 positions
            if sheet.cell(row,18).value == '20' or \
                sheet.cell(row,18).value == 20:

                long_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = long_files[-1].replace('\\','/').split('/')[-1].split('.')[0]


                file_info = long_files[-1].replace('\\','/').split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date

                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,18).value == 'xs' or \
                    sheet.cell(row,18).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' long' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' long' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.replace('\\','/').split('/')[-1]] = (long_files[-1].split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

            # Get Short 20 positions
            if sheet.cell(row,19).value == '20' or \
                sheet.cell(row,19).value == 20:

                short_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = short_files[-1].split('/')[-1].split('.')[0]
                file_info = short_files[-1].split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date
                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,19).value == 'xs' or \
                    sheet.cell(row,19).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.replace('\\','/').split('/')[-1]] = (short_files[-1].replace('\\','/').split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

            # Get Long and Short 20 positions
            if sheet.cell(row,20).value == '20' or \
                sheet.cell(row,20).value == 20:

                long_short_files.append(sheet.cell(row,2).value.split(' ')[0].replace('\\','/'))
                actual_file = sheet.cell(row,2).value.replace('\\','/')

                start_date = sheet.cell(row,3).value
                start_date_info = start_date.split('/')
                add_date = "_"+ start_date_info[1] +"_"+ short_month_names[start_date_info[0]] +"_20"+ start_date_info[2]

                end_date = sheet.cell(row,4).value
                end_date_info = end_date.split('/')
                folder_date_info = end_date.split('/')
                if len(folder_date_info[0]) == 1:
                    folder_date_info[0] = '0' + folder_date_info[0]
                if len(folder_date_info[1]) == 1:
                    folder_date_info[1] = '0' + folder_date_info[1]
                folder_date_info[2] = '20' + folder_date_info[2]
                current_output_date = folder_date_info[0] + folder_date_info[1] + folder_date_info[2]

                new_file = long_short_files[-1].split('/')[-1].split('.')[0]
                file_info = long_short_files[-1].split('/')[1].split(' ')[0]
                output_folder = 'Output-' + current_output_date
                if not os.path.isdir(output_date+'/'+output_folder):
                    os.makedirs(output_date+'/'+output_folder)
                if not output_folder in all_output_folders:
                    shutil.rmtree(output_date+'/'+output_folder)
                    os.makedirs(output_date+'/'+output_folder)
                    all_output_folders.append(output_folder)
                if not '2020' in new_file and not '2019' in new_file and not '2021' in new_file:
                    new_file += add_date


                if sheet.cell(row,20).value == 'xs' or \
                    sheet.cell(row,20).value == 'XS':
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' long+short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + " sticky.xls"
                else:
                    output_filename = output_date + '/' + output_folder + '/'+ new_file + '_top_20' + " " + horizon_info + ' long+short' + " until " + end_date_info[1] + " " + month_names[end_date_info[0]] + " 20" + end_date_info[2] + ".xls"

                true_dates[output_filename.split('/')[-1]] = (long_short_files[-1].split('/')[2])
                the_sheet_names[output_filename.replace('\\','/').split('/')[-1]] = sheet.cell(row,2).value.replace('\\','/')
                shutil.copyfile(actual_file,output_filename)

        index += 1

    print(all_output_folders)


    if not os.path.isdir(output_date+'/OutputFailed/'):
        os.makedirs(output_date+'/OutputFailed/')
    else:
        shutil.rmtree(output_date+'/OutputFailed/')
        os.makedirs(output_date+'/OutputFailed/')


    for i, output_folder in enumerate(all_output_folders):
        start_process(abs_date,output_date+'/'+output_folder,true_dates,the_sheet_names)
    # generate_xml(output_date+'/','https://iknowfirst.com.br/')
    # generate_xml(output_date+'/',SITE_URL)


export_files()













#
